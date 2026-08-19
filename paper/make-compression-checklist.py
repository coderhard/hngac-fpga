"""Build the copy-paste tracking workbook from the trim pairs.
Pulls the 43 (old, new) pairs straight out of apply-trim-2026-08-19.py so the
sheet can never drift from the script that was verified against the manuscript."""
import re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

REPO = '/home/abuhassan/projects/hngac-fpga'
TRIM = f'{REPO}/paper/apply-trim-2026-08-19.py'
AUTH = f'{REPO}/paper/main-author-rev-2026-08-19.tex'
OUT  = f'{REPO}/paper/compression-checklist-2026-08-19.xlsx'

# Pull R out of the trim script without running its file I/O or sys.exit.
ns = {}
code = open(TRIM).read().split("src = open(SRC).read()")[0]
exec(compile(code, TRIM, 'exec'), ns)
R = ns['R']

# (section, why) in the same order the rep() calls appear in the trim script.
META = [
 ("Roadmap (I)",      "Signposting, not argument"),
 ("II Related Work",  "105-word sentence, 4 semicolons -> 5 sentences"),
 ("II Related Work",  "Nested apposition; survey cite folded into cite group"),
 ("II Related Work",  "'we cite them as such' = self-congratulation"),
 ("II Related Work",  "Trailing-participle pileup (GenAI marker)"),
 ("II Related Work",  "60-word opener; claim before justification"),
 ("III-A Background", "Balanced-antithesis pose -> mechanism"),
 ("III-B Attack cls", "Trailing clause tightened"),
 ("III-B Attack cls", "Wordy subject"),
 ("III-B Attack cls", "FIX: punchline landed after the digression"),
 ("III-B Attack cls", "Redundant 'delivery property' gloss"),
 ("III-C Assumptions","'We assume the following.' is a heading, not a sentence"),
 ("IV Model",         "Retire one 'That is,'; cut 'One scope statement belongs here.'"),
 ("IV Model",         "Colon splice -> two sentences"),
 ("IV Model",         "Don't tell the reader what matters; cut 'precisely'"),
 ("V Architecture",   "Trailing clause inferable from prior sentence"),
 ("V Architecture",   "LARGEST CUT: paragraph re-narrated the figure caption"),
 ("V Architecture",   "'That is what we mean by' = glossing your own phrase"),
 ("V Architecture",   "Cut 'We note one design consequence plainly.'; drop epistrophe"),
 ("VI-A Platforms",   "Semicolon -> period"),
 ("VI-B Corpus",      "Interrupting apposition holds the verb 8 words"),
 ("VI-C Honesty",     "Cut the promise to be honest; keep the honesty"),
 ("VII-A Key finding","Merge three short sentences"),
 ("VII-A Key finding","Cut 'we draw attention to it rather than let a reader find it'"),
 ("VII-A Key finding","Cut 'we mean the second clause literally' (worst GenAI tell)"),
 ("VII-B Security",   "Wordy: 'silicon-bound RTL, not a software model of it'"),
 ("VII-B Security",   "Drop 'with respect to the question it can ask' scaffolding"),
 ("VII-B Security",   "FIX: 'Reading down a row' -> 'Across the total row'"),
 ("VII-B Security",   "'on facing evidence' is not an idiom"),
 ("VII-C Software",   "Cut 'the contrast is the paper in one image'"),
 ("VII-D Baselines",  "FIX: 'the April corpus' leaks an internal artifact name"),
 ("VII-E Adversarial","Three sentences to make a two-part distinction"),
 ("VII-E Adversarial","Trim 'architectural', 'strictly', 'measured'"),
 ("VII-F On silicon", "Comma splice"),
 ("VII-F On silicon", "Escalating tricolon -> checkable claim"),
 ("VII-F On silicon", "Cut the justification for reporting; keep the explanation"),
 ("VII-G Mean vs WC", "Heading should be indexable, not an essay title"),
 ("VII-G Mean vs WC", "Lead with the concession in 6 words, not 22"),
 ("VII-G Mean vs WC", "Colon splice; 'both concessions on one wall-clock axis'"),
 ("VII-G Mean vs WC", "Defensive throat-clearing before a supported claim"),
 ("VIII Limitations", "Cut opener; 3 duplications of text already in IV and VI"),
 ("IX Conclusion",    "'structure our next steps', 'precisely'"),
 ("IX Conclusion",    "100-word sentence, 3 comma splices; un-invert permit claim"),
]
assert len(META) == len(R), f"{len(META)} meta vs {len(R)} pairs"

auth = open(AUTH).read()
def lineno(s):
    return auth[:auth.index(s)].count('\n') + 1

def words(s):
    t = re.sub(r'\\[a-zA-Z]+\*?', '', s)
    t = re.sub(r'[{}]', ' ', t)
    return len(t.split())

rows = []
for i, ((old, new), (sec, why)) in enumerate(zip(R, META), start=1):
    rows.append({
        'n': i, 'line': lineno(old), 'sec': sec, 'why': why,
        'saved': words(old) - words(new),
        'old': old, 'new': new,
    })
rows.sort(key=lambda r: r['line'])
for i, r in enumerate(rows, start=1):
    r['n'] = i

wb = Workbook()

# ---------------- sheet 1: the checklist ----------------
ws = wb.active
ws.title = 'Edits'
HEAD = ['#', 'Done', 'Line', 'Section', 'Words saved', 'Why', 'FIND (old text)', 'REPLACE WITH (new text)']
ws.append(HEAD)

hfill = PatternFill('solid', fgColor='2F4858')
thin  = Side(style='thin', color='C9CDD1')
for c, _ in enumerate(HEAD, start=1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = hfill
    cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
ws.row_dimensions[1].height = 30

for r in rows:
    ws.append([r['n'], '', r['line'], r['sec'], r['saved'], r['why'], r['old'], r['new']])

nrows = ws.max_row
for row in range(2, nrows + 1):
    ws.row_dimensions[row].height = 96
    for col in range(1, len(HEAD) + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(vertical='top', wrap_text=(col >= 6))
    for col in (1, 2, 3, 5):
        ws.cell(row=row, column=col).alignment = Alignment(vertical='top', horizontal='center')
    ws.cell(row=row, column=4).alignment = Alignment(vertical='top', wrap_text=True)
    # monospace for the paste columns so LaTeX braces/backslashes read cleanly
    for col in (7, 8):
        ws.cell(row=row, column=col).font = Font(name='Consolas', size=9)

widths = {'A': 5, 'B': 8, 'C': 7, 'D': 18, 'E': 8, 'F': 34, 'G': 68, 'H': 68}
for k, v in widths.items():
    ws.column_dimensions[k].width = v

# tick-box column: dropdown + green strikethrough-ish highlight when done
dv = DataValidation(type='list', formula1='"DONE,SKIP"', allow_blank=True, showDropDown=False)
dv.prompt = 'Mark DONE once pasted into Overleaf'
dv.promptTitle = 'Status'
ws.add_data_validation(dv)
dv.add(f'B2:B{nrows}')

rng = f'A2:H{nrows}'
ws.conditional_formatting.add(rng, FormulaRule(
    formula=['$B2="DONE"'],
    fill=PatternFill('solid', fgColor='DFF0DA'), font=Font(color='6B7B6B')))
ws.conditional_formatting.add(rng, FormulaRule(
    formula=['$B2="SKIP"'],
    fill=PatternFill('solid', fgColor='F0F0F0'), font=Font(color='9A9A9A')))
# flag the three that are factual fixes, not style
ws.conditional_formatting.add(f'F2:F{nrows}', FormulaRule(
    formula=['AND($B2="",LEFT($F2,4)="FIX:")'],
    fill=PatternFill('solid', fgColor='FCE4E4'), font=Font(bold=True, color='A03030')))

ws.freeze_panes = 'C2'
ws.auto_filter.ref = f'A1:H{nrows}'

# ---------------- sheet 2: progress + notes ----------------
s2 = wb.create_sheet('Summary')
s2['A1'] = 'IPCCC 2026 compression pass — 9 pages to 8'
s2['A1'].font = Font(bold=True, size=14)
s2['A3'] = 'Source file'; s2['B3'] = 'paper/main-author-rev-2026-08-19.tex'
s2['A4'] = 'Verified result'; s2['B4'] = 'paper/main-trimmed-2026-08-19.tex (8 pages, 0 overfull boxes)'
s2['A5'] = 'Reproduce'; s2['B5'] = 'python3 paper/apply-trim-2026-08-19.py'
s2['A7'] = 'Edits total'; s2['B7'] = f'={nrows-1}'
s2['A8'] = 'Marked DONE'; s2['B8'] = f'=COUNTIF(Edits!B2:B{nrows},"DONE")'
s2['A9'] = 'Marked SKIP'; s2['B9'] = f'=COUNTIF(Edits!B2:B{nrows},"SKIP")'
s2['A10'] = 'Remaining'; s2['B10'] = f'=B7-B8-B9'
s2['A11'] = 'Words saved so far'; s2['B11'] = f'=SUMIF(Edits!B2:B{nrows},"DONE",Edits!E2:E{nrows})'
s2['A12'] = 'Words saved if all applied'; s2['B12'] = f'=SUM(Edits!E2:E{nrows})'
s2['A13'] = 'Words still needed for 8 pages'; s2['B13'] = '=MAX(0,500-B11)'
s2['A15'] = 'NOTE'
s2['B15'] = ('~500 words is roughly one IEEE column, which is the page-9 overflow. '
             'Rows 17-19 are factual fixes, not style: take those regardless of the page count.')
s2['B15'].alignment = Alignment(wrap_text=True, vertical='top')
s2['A17'] = 'Also pending (not in this sheet)'
s2['B17'] = ('1. New Fig. 1 scenario diagram (draw.io) -> figures/fig-scenario.pdf   '
             '2. Merge Tables II and IV into one "same cycles, different decisions" table   '
             '3. Add "a 19% increase" to VII-C (abstract cites it, body never states it)   '
             '4. Reconcile the 2,307 vs 200,000 request corpora in VI-B   '
             '5. Put the over-authorization count into Table V as a column')
s2['B17'].alignment = Alignment(wrap_text=True, vertical='top')
s2.row_dimensions[15].height = 46
s2.row_dimensions[17].height = 92
for r in (3,4,5,7,8,9,10,11,12,13,15,17):
    s2.cell(row=r, column=1).font = Font(bold=True)
s2.column_dimensions['A'].width = 30
s2.column_dimensions['B'].width = 86
for r in (8, 10, 11):
    s2.cell(row=r, column=2).font = Font(bold=True)

wb.save(OUT)
print(f"wrote {OUT}")
print(f"rows: {nrows-1}   total words saved: {sum(r['saved'] for r in rows)}")
print(f"factual fixes flagged: {sum(1 for r in rows if r['why'].startswith('FIX:'))}")
